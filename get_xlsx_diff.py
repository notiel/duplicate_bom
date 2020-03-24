# main script for comparing BOM's and finding duplicates and write resut to xlsx
# run with parameters «--compare filename1 filename2» to compare two BOMs by pns
#                     «--quantity filename1 filename2» to compare two BOMs by pns and quantity



import data_types
from typing import List, Optional, Dict, Union
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from string import ascii_uppercase
import sys
import xlsx_parce
import os
import compare_boms


def get_component_by_pn(pn: str, footprint: str, components=List[data_types.Component]) \
        -> Optional[data_types.Component]:
    """
    finds component by pn
    :param pn: pn of component
    :param footprint: footprint of component to find
    :param components: list of componentns
    :return: found component or None
    """
    for component in components:
        if component.pn == pn and component.footprint == footprint:
            return component
    return None


def get_component_by_value(value: Union[str, int], footprint: str, comp_type: data_types.ComponentType,
                           components=List[data_types.Component]) \
        -> Optional[data_types.Component]:
    """
    finds component by value and footprint (capacitor, resistor, inductor)
    :param comp_type: type of component
    :param value: value of component
    :param footprint: footprint of component to find
    :param components: list of componentns
    :return: found component or None
    """
    for component in components:
        if component.component_type == comp_type:
            if comp_type == data_types.ComponentType.CAPACITOR:
                if component.details:
                    if component.details.absolute_pf_value == value and component.footprint == footprint:
                        return component
            elif component.details and component.details.value == value and component.footprint == footprint:
                return component
    return None


def color_row(headers: Dict[str, Optional[int]], sheet, color: str, index: int):
    """
    makes row coloured
    :param index: row
    :param headers: headers of row
    :param sheet: sheet with data
    :param color: color
    :return:
    """
    for key in headers.keys():
        if headers[key]:
            sheet['%s%i' % (ascii_uppercase[headers[key] - 1], index)].fill = \
                PatternFill("solid", fgColor=color)


def add_row(component: data_types.Component, headers: Dict[str, Optional[int]], sheet, row: int, details: bool):
    """
    addes row to sheet
    :param details: add details for parametrized types
    :param row: row to add
    :param component: component to add
    :param headers: sheet headers
    :param sheet: sheet to add
    :return:
    """
    bd = Side(style='thin', color="000000")
    simple_keys = ['pn', 'manufacturer', 'footprint', 'description']
    if headers['type']:
        sheet["%s%i" % (ascii_uppercase[headers['type'] - 1], row)] = \
            data_types.types[component.component_type.value].capitalize()
    for key in simple_keys:
        if headers[key]:
            sheet["%s%i" % (ascii_uppercase[headers[key] - 1], row)] = getattr(component, key)
    if len(component.pn_alt):
        if headers['pn alternative 1']:
            sheet["%s%i" % (ascii_uppercase[headers['pn alternative 1'] - 1], row)] = component.pn_alt[0]
            if len(component.pn_alt) > 1:
                if headers['pn alternative 2']:
                    sheet["%s%i" % (ascii_uppercase[headers['pn alternative 2'] - 1], row)] = \
                        component.pn_alt[1]
    if headers['designator']:
        sheet["%s%i" % (ascii_uppercase[headers['designator'] - 1], row)] = ', '.join(component.designator)
    if headers['quantity']:
        sheet["%s%i" % (ascii_uppercase[headers['quantity'] - 1], row)] = len(component.designator)
    for key in headers.keys():
        if headers[key]:
            sheet['%s%i' % (ascii_uppercase[headers[key] - 1], row)].border = Border(left=bd, top=bd, right=bd,
                                                                                     bottom=bd)
    if details:
        add_details(component, headers, sheet, row)


def add_details(component: data_types.Component, headers: Dict[str, Optional[int]], sheet, row: int):
    """
    add detail data to sheet for resistor/inductor/capacitor
    :param component: component to add
    :param headers: sheet headers
    :param sheet: sheet to write data
    :param row: row to write data
    :return:
    """
    if headers['value']:
        if component.component_type == data_types.ComponentType.INDUCTOR:
            sheet["%s%i" % (ascii_uppercase[headers['value'] - 1], row)] = component.details.value
        elif component.component_type == data_types.ComponentType.RESISTOR:
            sheet["%s%i" % (ascii_uppercase[headers['value'] - 1], row)] = str(component.details.value) + 'R'
        else:
            sheet["%s%i" % (ascii_uppercase[headers['value'] - 1], row)] = \
                str(component.details.value) + str(data_types.units_cap[component.details.unit])
        if component.component_type in [data_types.ComponentType.RESISTOR, data_types.ComponentType.CAPACITOR]:
            if headers['tolerance']:
                sheet["%s%i" % (ascii_uppercase[headers['tolerance'] - 1], row)] = \
                    str(100 * component.details.tolerance) + '%'
        if component.component_type == data_types.ComponentType.CAPACITOR:
            if headers['voltage']:
                sheet["%s%i" % (ascii_uppercase[headers['voltage'] - 1], row)] = str(component.details.voltage) + 'V'
            if headers['dielectric']:
                sheet["%s%i" % (ascii_uppercase[headers['dielectric'] - 1], row)] = \
                    ', '.join([dielectric.name for dielectric in component.details.dielectric])


def find_new_pns(old: List[data_types.Component], new: List[data_types.Component], sheet,
                 headers: Dict[str, Optional[int]], quantity: bool):
    """
    compare two lists to find new positions
    :param quantity: add quantity?
    :param sheet: sheet
    :param headers: headers
    :param old: list of old components
    :param new: list of new components
    :return:
    """
    old_pn, old_cap, old_res, old_ind = compare_boms.get_comp_list_precise(old)
    new_pn, new_cap, new_res, new_ind = compare_boms.get_comp_list_precise(new)
    plus_pn, minus_pn, _ = compare_boms.get_diff_data(old_pn, new_pn, "partnumbers", True)
    plus_cap, minus_cap, _ = compare_boms.get_diff_data(old_cap, new_cap, "capacitors", True)
    plus_res, minus_res, _ = compare_boms.get_diff_data(old_res, new_res, "resistors", True)
    plus_ind, minus_ind, _ = compare_boms.get_diff_data(old_ind, new_ind, "inductors", True)
    new_sorted = sorted(new, key=lambda x: x.row)
    index = 0
    for (index, component) in enumerate(new_sorted):
        if component.component_type not in [data_types.ComponentType.CAPACITOR, data_types.ComponentType.RESISTOR]:
            if component.pn:
                add_row(component, headers, sheet, index + 2, False)
            if quantity:
                paired = get_component_by_pn(component.pn, component.footprint, old)
                if paired:
                    if len(paired.designator) > len(component.designator):
                        sheet["%s%i" % (ascii_uppercase[headers['quantity'] - 1], index + 2)].fill = \
                            PatternFill("solid", fgColor='FF0000')
                    elif len(paired.designator) < len(component.designator):
                        sheet["%s%i" % (ascii_uppercase[headers['quantity'] - 1], index + 2)].fill = \
                            PatternFill("solid", fgColor="00FF00")
            if (component.pn, component.footprint) in plus_pn:
                color_row(headers, sheet, '00FF00', index + 2)

        if component.component_type == data_types.ComponentType.RESISTOR:
            if component.details and component.details.value != -1:
                add_row(component, headers, sheet, index + 2, True)
                if (str(component.details.value) + 'R', component.footprint) in plus_res:
                    color_row(headers, sheet, '00FF00', index + 2)
                if quantity:
                    paired = get_component_by_value(str(component.details.value) + 'R',
                                                    component.footprint, component.component_type, old)
                    if paired:
                        if len(paired.designator) > len(component.designator):
                            sheet["%s%i" % (ascii_uppercase[headers['quantity'] - 1], index + 2)].fill = \
                                PatternFill("solid", fgColor='FF0000')
                        elif len(paired.designator) < len(component.designator):
                            sheet["%s%i" % (ascii_uppercase[headers['quantity'] - 1], index + 2)].fill = \
                                PatternFill("solid", fgColor="00FF00")

        if component.component_type == data_types.ComponentType.CAPACITOR:
            if component.details and component.details.value:
                add_row(component, headers, sheet, index + 2, True)
                if (str(component.details.absolute_pf_value) + 'pf', component.footprint) in plus_cap:
                    color_row(headers, sheet, '00FF00', index + 2)

                if quantity:
                    paired = get_component_by_value(str(component.details.absolute_pf_value) + 'pf',
                                                    component.footprint, component.component_type, old)
                    if paired:
                        if len(paired.designator) > len(component.designator):
                            sheet["%s%i" % (ascii_uppercase[headers['quantity'] - 1], index + 2)].fill = \
                                PatternFill("solid", fgColor='FF0000')
                        elif len(paired.designator) < len(component.designator):
                            sheet["%s%i" % (ascii_uppercase[headers['quantity'] - 1], index + 2)].fill = \
                                PatternFill("solid", fgColor="00FF00")

        if component.component_type == data_types.ComponentType.INDUCTOR:
            if not component.pn and component.details and component.details.value:
                add_row(component, headers, sheet, index + 2, True)
                if (component.details.value, component.footprint) in plus_ind:
                    color_row(headers, sheet, '00FF00', index + 2)
                if quantity:
                    paired = get_component_by_value(str(component.details.value),
                                                    component.footprint, component.component_type, old)
                    if paired:
                        if len(paired.designator) > len(component.designator):
                            sheet["%s%i" % (ascii_uppercase[headers['quantity'] - 1], index + 2)].fill = \
                                PatternFill("solid", fgColor='FF0000')
                        elif len(paired.designator) < len(component.designator):
                            sheet["%s%i" % (ascii_uppercase[headers['quantity'] - 1], index + 2)].fill = \
                                PatternFill("solid", fgColor="00FF00")

    added_rows = index + 3
    for (pn, footprint) in minus_pn:
        component = get_component_by_pn(pn, footprint, old)
        if component:
            add_row(component, headers, sheet, added_rows, False)
            color_row(headers, sheet, 'FF0000', added_rows)
            added_rows += 1
    for (value, footprint) in minus_cap:
        component = get_component_by_value(value, footprint, data_types.ComponentType.CAPACITOR, old)
        if component:
            add_row(component, headers, sheet, added_rows, False)
            color_row(headers, sheet, 'FF0000', added_rows)
            added_rows += 1
    for (value, footprint) in minus_res & minus_ind:
        component = get_component_by_value(value, footprint, data_types.ComponentType.RESISTOR, old)
        if component:
            add_row(component, headers, sheet, added_rows, False)
            color_row(headers, sheet, 'FF0000', added_rows)
            added_rows += 1


def write_results(old: List[data_types.Component], new: List[data_types.Component], filename1: str, filename2: str,
                  headers: Dict[str, Optional[int]], quantity=False):
    """
    write results to results.xlsx workbook
    :param quantity: use quantity
    :param new: list with new components
    :param headers: headers
    :param filename2: name of second file
    :param filename1: name of first file
    :param old: list with components
    :return:
    """
    wb = Workbook()
    ws1 = wb.active
    for header in headers.keys():
        if headers[header]:
            ws1['%s1' % ascii_uppercase[headers[header] - 1]] = header.capitalize()
            ws1['%s1' % ascii_uppercase[headers[header] - 1]].font = Font(b=True)
            ws1['%s1' % ascii_uppercase[headers[header] - 1]].alignment = Alignment(horizontal="center")
            ws1.column_dimensions[ascii_uppercase[headers[header] - 1]].width = 17
    find_new_pns(old, new, ws1, headers, quantity)
    filename1 = os.path.basename(filename1)
    filename2 = os.path.basename(filename2)
    wb.save(filename=os.path.splitext(filename1)[0] + '_' + os.path.splitext(filename2)[0] + 'diff.xlsx')


def compare_boms_new_pns(quantity=False, detailed=False):
    """
    main function for bom comparing. Use detailed=True to get detailed comparing and False to compare PNs only
    :return:
    """
    if len(sys.argv) != 4:
        print("Wrong params: two bom files to compare expected")
        return
    if not os.path.exists(sys.argv[2]) or not os.path.exists(sys.argv[3]) or os.path.isdir(sys.argv[2]) \
            or os.path.isdir(sys.argv[3]):
        print("Both files should be existing files (not folders)")
        return
    old, warning = xlsx_parce.get_components_from_xlxs(sys.argv[2])
    print(warning)
    new, warning = xlsx_parce.get_components_from_xlxs(sys.argv[3])
    print(warning)
    write_results(old, new, sys.argv[2], sys.argv[3],
                  xlsx_parce.get_headers(load_workbook(filename=sys.argv[3]).active), quantity)
    if detailed:
        compare_boms.detail_compare(old, new, False)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Parameters are missing, need type parameters and 1 or 2 filenames")
    if sys.argv[1].lower() == '--compare':
        compare_boms_new_pns()
    elif sys.argv[1].lower() == '--detailed':
        compare_boms_new_pns(detailed=True)
    elif sys.argv[1].lower() == '--quantity':
        compare_boms_new_pns(quantity=True)
    else:
        print("Wrong parameter")
