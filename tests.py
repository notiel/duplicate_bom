import unittest
import xlsx_parce
import duplicates
from openpyxl import load_workbook
import data_types
import compare_boms

header_index_etalon = {'type': 1, 'pn': 2, 'manufacturer': 3, 'pn alternative 1': 4, 'pn alternative 2': 5,
                       'designator': 6, 'footprint': 7, 'dielectric': 8, 'value': 9, 'voltage': 10, 'tolerance': 11,
                       'description': 12, 'ref': None, 'reference': None}

pn_diff_etalon = """
Added partnumbers in second file:{('BAT54HT1G', 'SOD323')} 
Deleted partnumbers in second file:{{('KX-6 37.4 MHz 10/10ppm 18pF', '2520'), ('U.FL-R-SMT-1', 'U.FL-R-SMT-1'), 
('AP6398S', 'AP6356S'), ('WPM3401', 'SOT23'), ('LMBR160FT1G', 'SOD123-FL')}}"""

filename = "test_data\\v2.1.xlsx"
filename_new = "test_data\\v2.1.xlsx"
filename_old = 'test_data\\V1.6.xlsx'
filename_duplicate = 'test_data\\BOM_Test.xlsx'

types_etalon = {1: 14, 5: 9, 8: 1, 2: 2, 4: 2, 7: 3, 3: 3, 0: 30, 11: 1}


class XLSXParceTest(unittest.TestCase):

    def setUp(self):
        self.data, self.warning = xlsx_parce.get_components_from_xlxs(filename)

    def testRowNumber(self):
        self.assertEqual(len(self.data), 65)

    def testNotUsed(self):
        self.assertEqual(self.warning.count('not used'), 3)

    def testHeader(self):
        wb = load_workbook(filename=filename)
        sheet = wb.active
        header_index = xlsx_parce.get_headers(sheet)
        self.assertEqual(header_index, header_index_etalon)

    def testTypesCorrect(self):
        types = dict()
        for component in self.data:
            if component.component_type.value in types.keys():
                types[component.component_type.value] += 1
            else:
                types[component.component_type.value] = 1
        self.assertEqual(types_etalon, types)

    def testResistorData(self):
        wb = load_workbook(filename=filename)
        sheet = wb.active
        self.assertEqual(xlsx_parce.get_resistor_data((sheet, 43, header_index_etalon)),
                         data_types.Resistor(value=1000, tolerance=0.01))

    def testCapacitorData(self):
        wb = load_workbook(filename=filename)
        sheet = wb.active
        self.assertEqual(xlsx_parce.get_capacitor_data((sheet, 16, header_index_etalon)),
                         data_types.Capacitor(value=22, unit=data_types.CapUnits.U, absolute_pf_value=22000000,
                                              dielectric=[data_types.Dielectric.X5R, data_types.Dielectric.X7R],
                                              voltage=6.3, tolerance=0.2))

    def testNotUsedItem(self):
        wb = load_workbook(filename=filename)
        sheet = wb.active
        self.assertTrue(xlsx_parce.check_for_not_used((sheet, 14, header_index_etalon)))
        self.assertFalse(xlsx_parce.check_for_not_used((sheet, 15, header_index_etalon)))


class FootprintsTest(unittest.TestCase):

    def testFootPrint(self):
        self.assertEqual(xlsx_parce.get_footprint_data("C0805_YA", data_types.ComponentType.CAPACITOR), '0805')

    def testComponentTypeyFootprint(self):
        self.assertEqual(xlsx_parce.get_component_type_by_footprint("Capacitors:CAP_0603"),
                         data_types.ComponentType.CAPACITOR)

    def testCrystalFootpint(self):
        self.assertEqual(xlsx_parce.get_footprint_data('CRY_2520_4P', data_types.ComponentType.CRYSTAL), '2520')


class ResistorTest(unittest.TestCase):

    def testIncorrectValue(self):
        self.assertIsNone(xlsx_parce.get_resistor_value("wrong data"))

    def testFraction(self):
        self.assertEqual(xlsx_parce.get_resistor_value('0r075'), 0.075)

    def testFraction2(self):
        self.assertEqual(xlsx_parce.get_resistor_value('0.01R'), 0.01)

    def testZero(self):
        self.assertEqual(xlsx_parce.get_resistor_value('0'), 0)

    def testKValue(self):
        self.assertEqual(xlsx_parce.get_resistor_value('4k7'), 4700)

    def testKValue2(self):
        self.assertEqual(xlsx_parce.get_resistor_value('5.1k'), 5100)

    def testMValue(self):
        self.assertEqual(xlsx_parce.get_resistor_value('1M'), 1000000)


class CapacitorDataTest(unittest.TestCase):

    def testNf(self):
        self.assertEqual(xlsx_parce.get_capacitor_value_and_unit("100nF"), (100, data_types.units_cap.index('n'),
                                                                            [data_types.Dielectric.X5R,
                                                                             data_types.Dielectric.X7R], 100000))

    def testU(self):
        self.assertEqual(xlsx_parce.get_capacitor_value_and_unit("4.7uF"), (4.7, data_types.units_cap.index('u'),
                                                                            [data_types.Dielectric.X5R,
                                                                             data_types.Dielectric.X7R], 4700000))

    def testPf(self):
        self.assertEqual(xlsx_parce.get_capacitor_value_and_unit("6.8pF"), (6.8, data_types.units_cap.index('pf'),
                                                                            [data_types.Dielectric.NP0], 6.8))

    def testError(self):
        self.assertEqual(xlsx_parce.get_capacitor_value_and_unit("4.7fF"), (None, None, None, None))


class CompareTest(unittest.TestCase):

    def setUp(self):
        self.data, _ = xlsx_parce.get_components_from_xlxs(filename_duplicate)

    def testEqualPns(self):
        equal, _, _, _ = duplicates.compare_pns(self.data, 8, False)
        self.assertEqual(equal, [(filename_duplicate.split('\\')[1], 17, 13,
                                  filename_duplicate.split('\\')[1], 35, 31)])

    def testSimilarPns(self):
        _, similar, _, _ = duplicates.compare_pns(self.data, 5, False)
        self.assertEqual(similar, [(filename_duplicate.split('\\')[1], 30, 9,
                                    filename_duplicate.split('\\')[1], 31, 10),
                                   (filename_duplicate.split('\\')[1], 30, 9,
                                    filename_duplicate.split('\\')[1], 32, 11),
                                   (filename_duplicate.split('\\')[1], 25, 17,
                                    filename_duplicate.split('\\')[1], 26, 18)])

    def testAltPns(self):
        _, _, alternative, _ = duplicates.compare_pns(self.data, 8, False)
        self.assertEqual(alternative, [(filename_duplicate.split('\\')[1], 19, 1,
                                        filename_duplicate.split('\\')[1], 23, 19),
                                       (filename_duplicate.split('\\')[1], 30, 9,
                                        filename_duplicate.split('\\')[1], 31, 27)])

    def testPNPrecise(self):
        equal, similar, alt, _ = duplicates.compare_pns(self.data, 8, True)
        self.assertEqual(equal,
                         [(filename_duplicate.split('\\')[1], 17, 13, filename_duplicate.split('\\')[1], 35, 31)])
        self.assertEqual(similar, list())
        self.assertEqual(alt, list())

    def testCompareCapacitors(self):
        similar_cap = duplicates.compare_capacitors(self.data)
        self.assertEqual(similar_cap, [(filename_duplicate.split('\\')[1], '0.01n', 7,
                                        filename_duplicate.split('\\')[1], '10pf', 12),
                                       (filename_duplicate.split('\\')[1], '0.1u', 2,
                                        filename_duplicate.split('\\')[1], '100n', 5)])

    def testCompareResistors(self):
        similar_res = duplicates.compare_resistors(self.data)
        self.assertEqual(similar_res, [(filename_duplicate.split('\\')[1], '4700R', 52,
                                        filename_duplicate.split('\\')[1], '4700.0R', 53),
                                       (filename_duplicate.split('\\')[1], '22000R', 55,
                                        filename_duplicate.split('\\')[1], '22000R', 56)])


class CompareBoms(unittest.TestCase):

    def setUp(self):
        self.new, _ = xlsx_parce.get_components_from_xlxs(filename_new)
        self.old, _ = xlsx_parce.get_components_from_xlxs(filename_old)
        self.old_pn, self.old_cap, self.old_res, self.old_ind = compare_boms.get_comp_list_precise(self.old)
        self.new_pn, self.new_cap, self.new_res, self.new_ind = compare_boms.get_comp_list_precise(self.new)

    def testPNDiff(self):
        plus, minus, _ = compare_boms.get_diff_data(self.old_pn, self.new_pn, "partnumbers", True)
        self.assertEqual(plus, {('LMBR160FT1G', 'SOD123-FL')})
        self.assertEqual(minus, {('AP6398S', 'AP6356S'),
                                 ('KX-6 37.4 MHz 10/10ppm 18pF', '2520'),
                                 ('NU/LMBR160FT1G', 'SOD123-FL'),
                                 ('U.FL-R-SMT-1', 'U.FL-R-SMT-1'),
                                 ('WPM3401', 'SOT23')})

    def testCapDiff(self):
        plus, minus, _ = compare_boms.get_diff_data(self.old_cap, self.new_cap, "capacitors", True)
        self.assertEqual(plus, {('18pf', '0603')})
        self.assertEqual(minus, {('10pf', '0402'), ('18pf', '0402')})

    def testResDiff(self):
        plus, minus, _ = compare_boms.get_diff_data(self.old_res, self.new_res, "resistors", True)
        self.assertEqual(plus, {('12000R', '0402')})
        self.assertEqual(minus, {('127000R', '0402')})

    def testNoDeleted(self):
        _, _, res = compare_boms.get_diff_data(self.old_pn, self.new_pn, "partnumbers", False)
        self.assertFalse('deleted' in res.lower())

    def testComponentsInList(self):
        self.assertEqual(self.new[3], compare_boms.find_components_in_list(self.old[3], self.new))

    def testAbsentComponents(self):
        component = [component for component in self.old if component.pn == 'AP6398S'][0]
        self.assertIsNone(compare_boms.find_components_in_list(component, self.new))


class TestJoin(unittest.TestCase):

    def setUp(self):
        self.new, _ = xlsx_parce.get_components_from_xlxs(filename_new)
        self.old, _ = xlsx_parce.get_components_from_xlxs(filename_old)
        self.old_len = len(self.old)
        compare_boms.join_the_same(self.old)

    def testJoinPN(self):
        self.assertEqual(len(self.old) + 3, self.old_len)
        joined_component = [component for component in self.old if component.pn == 'TLV62569PDDC'][0]
        self.assertEqual(len(joined_component.designator), 4)

    def testJoinCap(self):
        joined_cap = [component for component in self.old if component.row == 8][0]
        self.assertEqual(len(joined_cap.designator), 8)
        self.assertTrue('C250' in joined_cap.designator)

    def testJoinRes(self):
        joined_res = [component for component in self.old if component.row == 49][0]
        self.assertEqual(len(joined_res.designator), 5)
        self.assertTrue('R200' in joined_res.designator)

    # def testCorrect(self):
    #    component_false = [component for component in self.old if component.row == 46][0]
    #    self.assertFalse(compare_boms.check_component(component_false))

    def tearUp(self):
        self.new = self.old = self.old_pn = self.old_cap = self.old.res = self.old_ind = self.new_cap = list()
        self.new_pn = self.new_res = self.new_ind = list()


class TestPNDescription(unittest.TestCase):

    def setUp(self):
        self.new, _ = xlsx_parce.get_components_from_xlxs(filename_new)
        self.old, _ = xlsx_parce.get_components_from_xlxs(filename_old)

    def testPNDescription(self):
        self.assertEqual(compare_boms.get_pn(self.new[16]), 'PN: KLM8G1GETF-B041, footprint: BGA153')

    def testCapDescription(self):
        self.assertEqual(compare_boms.get_pn(self.new[5]), "CAPACITOR with value 1u, footprint: 0402")

    def testResDescription(self):
        self.assertEqual(compare_boms.get_pn(self.new[60]), 'RESISTOR with value 240R, footprint: 0402')

    def testInsuctor(self):
        self.assertEqual(compare_boms.get_pn(self.new[30]), 'PN: LQH3NPN4R7MGR, footprint: LQH3NPN')


if __name__ == '__main__':
    unittest.main()
