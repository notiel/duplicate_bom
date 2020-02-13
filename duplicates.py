# from openpyxl import Workbook
from openpyxl import load_workbook
from typing import List, Optional, Union, Tuple
from string import ascii_uppercase

import data_types

headers = ['type', 'pn', 'manufacturer', 'pn alternative 1', 'pn alternative 2', 'designator', 'footprint',
           'dielectric', 'value', 'voltage', 'tolerance', 'description']
recommended = ['type', 'pn', 'designator', 'footprint', 'value']
multiplier = {'m': 6, 'k': 3, 'r': 0}
warning = ""


def get_value(key: str, row: int, sheet, header_index) -> str:
    """
    gets value from sheet row by the key
    :param row: row of value
    :param key: key of header
    :param sheet: sheet with data
    :param header_index: list of header indexes
    :return: value
    """
    if header_index[key] and sheet.cell(row=row, column=header_index[key]):
        return sheet.cell(row=row, column=header_index[key]).value
    return ""


def get_component_type(comp_type_str: str) -> data_types.ComponentType:
    """
    gets component type analazing string
    :param comp_type_str: field type from xlsx
    :return: type of component
    """
    for comp_type in data_types.types:
        if comp_type.lower() in comp_type_str.lower():
            return data_types.ComponentType(data_types.types.index(comp_type))
    return data_types.ComponentType.OTHER


def get_resistor_value(resistor_str: str) -> Optional[Union[float, int]]:
    """
    converts str value of fesistor to float
    4k7 -> 4700
    100k -> 100000
    5r2 -> 5.2
    0r01 -> 0.01
    1m -> 1000000
    :param resistor_str:
    :return:
    """
    if isinstance(resistor_str, float) or isinstance(resistor_str, int):
        return resistor_str
    resistor_str = resistor_str.replace(',', '.').lower()
    convert = float if '.' in resistor_str else int
    for key in multiplier.keys():
        if key in resistor_str:
            try:
                index = resistor_str.index(key)
                tail = len(resistor_str) - index
                value = convert(resistor_str[:index]+resistor_str[index+1:])
                return value*10**(multiplier[key]-tail+1)
            except ValueError:
                return None
    try:
        return convert(resistor_str)
    except ValueError:
        return None


def get_capacitor_value_and_unit(capacitor_str:  str) -> Tuple[Optional[Union[int, float]],
                                                               Optional[data_types.CapUnits],
                                                               Optional[List[data_types.Dielectric]]]:
    """

    :param capacitor_str:
    :return:
    """
    capacitor_str = capacitor_str.replace(',', '.').lower()
    convert = float if '.' in capacitor_str else int
    if 'pf' in capacitor_str:
        try:
            value = convert(capacitor_str.replace("pf", ""))
            return value, data_types.CapUnits.PF, [data_types.Dielectric.NP0]
        except ValueError:
            return None, None, None
    capacitor_str = capacitor_str.replace('µ', 'u')
    capacitor_str = capacitor_str.replace('f', '')
    if 'u' in capacitor_str:
        try:
            value = convert(capacitor_str.replace("u", ""))
            return value, data_types.CapUnits.U, [data_types.Dielectric.X5R, data_types.Dielectric.X7R]
        except ValueError:
            return None, None, None
    elif 'n' in capacitor_str:
        try:
            value = convert(capacitor_str.replace("n", ""))
            return value, data_types.CapUnits.N, [data_types.Dielectric.X5R, data_types.Dielectric.X7R]
        except ValueError:
            return None, None, None
    return None, None, None


def get_components_from_xlxs(filename) -> List[data_types.Component]:
    """
    parce Bom to get component list
    :param filename: name of BOM
    :return: component list
    """
    wb = load_workbook(filename=filename)
    sheet = wb.active
    result: List[data_types.Component] = list()
    header_index = dict.fromkeys(headers, None)
    for col in ascii_uppercase:
        col_header = sheet['%s1' % col].value
        if col_header and col_header.lower() in headers:
            header_index[col_header.lower()] = ascii_uppercase.index(col) + 1
    absent_headers = [header for header in recommended if not header_index[header]]
    if absent_headers:
        global warning
        warning = 'The following columns are recommended but absent: ' + ', '.join(absent_headers)
    for row in range(1, sheet.max_row):
        comp_type = get_component_type(get_value('type', row, sheet, header_index))
        component = data_types.Component(component_type=comp_type,
                                         footprint=get_value('footprint', row, sheet, header_index),
                                         pn=get_value('pn', row, sheet, header_index),
                                         manufacturer=get_value('manufacturer', row, sheet, header_index),
                                         designator=get_value('designator', row, sheet, header_index).split(', '),
                                         description=get_value('description', row, sheet, header_index))
        if comp_type == data_types.ComponentType.RESISTOR:
            value = get_resistor_value(get_value('value', row, sheet, header_index))
            if not value:
                warning += 'Wrong resistor value at %i row\n' % row
            resistor = data_types.Resistor(value=value, tolerance=get_value('tolerance', row, sheet, header_index))
            component.details = resistor
        elif comp_type == data_types.ComponentType.CAPACITOR:
            value, unit, dielectric = get_capacitor_value_and_unit(get_value('value', row, sheet, header_index))
            if not value:
                warning += 'Wrong capacitor value at %i row\n' % row
            dielectric_bom = get_value('dielectric', row, sheet, header_index)
            if dielectric_bom:
                if not dielectric_bom.lower() in data_types.dielectrics\
                        or data_types.Dielectric(data_types.dielectrics.index(dielectric_bom.lower()) )not in dielectric:
                    warning += 'Dielectric from BOM does not match capacitor value in %i row\n' % row
            capacitor = data_types.Capacitor(value=value,
                                             unit=unit,
                                             voltage=get_value('voltage', row, sheet, header_index),
                                             tolerance=get_value('tolerance', row, sheet, header_index),
                                             dielectric=dielectric)
            print(capacitor)
            component.details = capacitor
        elif comp_type == data_types.ComponentType.INDUCTOR:
            inductor = data_types.Inductor(value=get_value('value', row, sheet, header_index))
            print(inductor)
            component.details = inductor
        result.append(component)
    return result


get_components_from_xlxs('BOM_StPlus_SOM_V1.5.xlsx')
print(warning)
