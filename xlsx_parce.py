# service module, xlsx BOM file parcing for find_duplicates

from openpyxl import Workbook
from openpyxl import load_workbook, worksheet
from typing import List, Optional, Union, Tuple, Any, Dict
from string import ascii_uppercase
import os

import data_types

headers = ['type', 'pn', 'manufacturer', 'pn alternative 1', 'pn alternative 2', 'designator', 'footprint',
           'dielectric', 'value', 'voltage', 'tolerance', 'description']
recommended = ['type', 'pn', 'designator', 'footprint', 'value']
multiplier = {'m': 6, 'k': 3, 'r': 0}
warning = ""

Row = Tuple[Any, int, Dict[str, Optional[Any]]]


def get_value(key: str, sheet: worksheet, row: int,  header_index: Dict[str, Optional[int]]) -> str:
    """
    gets value from sheet row by the key
    :param row: row of value
    :param key: key of header
    :param sheet: sheet with data
    :param header_index: list of header indexes
    :return: value
    """
    if header_index[key] and sheet.cell(row=row, column=header_index[key]):
        return sheet.cell(row=row, column=header_index[key]).value
    return ""


def get_component_type(comp_type_str: str) -> data_types.ComponentType:
    """
    gets component type analyzing string
    :param comp_type_str: field type from xlsx
    :return: type of component
    """
    if not comp_type_str:
        return data_types.ComponentType.OTHER
    for comp_type in data_types.types:
        if comp_type.lower() in comp_type_str.lower():
            return data_types.ComponentType(data_types.types.index(comp_type))
    return data_types.ComponentType.OTHER


def get_resistor_value(resistor_str: str) -> Optional[Union[float, int]]:
    """
    converts str value of fesistor to float
    4k7 -> 4700
    100k -> 100000
    5r2 -> 5.2
    0r01 -> 0.01
    1m -> 1000000
    :param resistor_str:
    :return:
    """
    if not resistor_str:
        return None
    # if resistor_str is already a digital value return itself
    if isinstance(resistor_str, float) or isinstance(resistor_str, int):
        return resistor_str
    resistor_str = resistor_str.replace(',', '.').lower()
    # we have float value for str with . and int for others
    convert = float if '.' in resistor_str else int
    for key in multiplier.keys():
        if key in resistor_str:
            try:
                # get a digital value,  multiply it according unit letter (r, k, m) and it place in string (47r != 4r7)
                index = resistor_str.index(key)
                tail = len(resistor_str) - index
                value = convert(resistor_str[:index] + resistor_str[index + 1:])
                return value * 10 ** (multiplier[key] - tail + 1)
            except ValueError:
                return None
    try:
        return convert(resistor_str)
    except ValueError:
        return None


def get_capacitor_value_and_unit(capacitor_str: str) -> Tuple[Optional[Union[int, float]],
                                                              Optional[data_types.CapUnits],
                                                              Optional[List[data_types.Dielectric]],
                                                              Optional[Union[int, float]]]:
    """
    gets capacitor value, unit and dielectric
    unit is nf, uf (micro) or pf
    dielectric is X5R or X7R for uf and nf and np0 for pf
    :param capacitor_str: string with capacitor description
    :return: value, unit and dielectric or None, absolute value in pfs
    """
    capacitor_str = capacitor_str.replace(',', '.').lower()
    # we have float value for str with . and int for others
    convert = float if '.' in capacitor_str else int
    try:
        if 'pf' in capacitor_str:
            value = convert(capacitor_str.replace("pf", ""))
            return value, data_types.CapUnits.PF, [data_types.Dielectric.NP0], value
        capacitor_str = capacitor_str.replace('µ', 'u')
        capacitor_str = capacitor_str.replace('f', '')
        if 'u' in capacitor_str:
            value = convert(capacitor_str.replace("u", ""))
            return value, data_types.CapUnits.U, [data_types.Dielectric.X5R, data_types.Dielectric.X7R], 10**6*value
        if 'n' in capacitor_str:
            value = convert(capacitor_str.replace("n", ""))
            return value, data_types.CapUnits.N, [data_types.Dielectric.X5R, data_types.Dielectric.X7R], 10**3*value
    except ValueError:
        return None, None, None, None
    return None, None, None, None


def get_footprint_data(footprint_str: str, comp_type: data_types.ComponentType) -> str:
    """
    deletes unnecessary data from footprint
    :param comp_type: type of component
    :param footprint_str: footprint from BOM
    :return: footprint without unnesessaty data
    """
    if not footprint_str:
        return ""
    first_letter: Dict[data_types.ComponentType, str] = {data_types.ComponentType.CAPACITOR: 'c',
                                                         data_types.ComponentType.INDUCTOR: 'i',
                                                         data_types.ComponentType.RESISTOR: 'r'}
    # remove type letter for capacitors/inductors/resistors and turn first part before _
    for key in first_letter.keys():
        if comp_type == key:
            if footprint_str.lower().startswith(first_letter[key]):
                return footprint_str[1:].split('_')[0]
    if comp_type is data_types.ComponentType.CRYSTAL:
        if footprint_str.lower().startswith('cry'):
            return footprint_str[4:]
    return footprint_str


def get_headers(sheet: Workbook) -> Dict[str, Optional[int]]:
    """
    get header indexes
    :param sheet: sheet with data
    :return:
    """
    header_index: Dict[str, Optional[str]] = dict.fromkeys(headers, None)
    for col in ascii_uppercase:
        col_header: str = sheet['%s1' % col].value
        if col_header and col_header.lower() in headers:
            header_index[col_header.lower()] = ascii_uppercase.index(col) + 1
    absent_headers: List[str] = [header for header in recommended if not header_index[header]]
    if absent_headers:
        global warning
        warning = 'The following columns are recommended but absent: ' + ', '.join(absent_headers)
    return header_index


def get_resistor_data(row_addr: Row) -> data_types.Resistor:
    """
    gets resistor data from sheet
    :param row_addr:
    :return: resistor instance
    """
    value:  Optional[Union[float, int]] = get_resistor_value(get_value('value', *row_addr))
    if not value:
        value = -1
    resistor = data_types.Resistor(value=value, tolerance=get_value('tolerance', *row_addr))
    return resistor


def get_capacitor_data(row_addr: Row) -> data_types.Capacitor:
    """
    get capacitor data from bom row
    :param row_addr: row address
    :return:
    """
    value, unit, dielectric, abs_value = get_capacitor_value_and_unit(get_value('value', *row_addr))
    if not value:
        value = 0
    dielectric_bom = get_value('dielectric', *row_addr)
    if dielectric_bom:
        for new_dielectric in dielectric_bom.replace("or ", "").lower().split():
            if new_dielectric not in data_types.dielectrics or\
                    data_types.Dielectric(data_types.dielectrics.index(new_dielectric)) not in dielectric:
                global warning
                warning += 'Dielectric from BOM does not match capacitor value in %i row\n' % row_addr[1]
    try:
        voltage = get_value('voltage', *row_addr).lower().replace("v", "")
    except (ValueError, AttributeError):
        voltage = 0
    try:
        tolerance = float(get_value('tolerance', *row_addr))
    except (ValueError, AttributeError, TypeError):
        tolerance = 0
    capacitor = data_types.Capacitor(value=value, unit=unit, voltage=voltage, tolerance=tolerance,
                                     dielectric=dielectric, absolute_pf_value=abs_value)
    return capacitor


def validate_and_repair(component: data_types.Component):
    """
    validates component, prints warning, returns corrected component
    :param component: component to validate
    :return: corrected component
    """
    errors: str = ""
    if component.component_type == data_types.ComponentType.OTHER:
        errors += "Unknown type\n"
    if not component.designator:
        errors += 'No designators\n'
        component.designator = list()
    if not component.description:
        component.description = ""
    if not component.pn_alt:
        component.pn_alt = list()
    if not component.manufacturer:
        component.manufacturer = ""
    if not component.footprint:
        errors += "No footprint\n"
        component.footprint = ""
    if component.component_type == data_types.ComponentType.CAPACITOR:
        if not component.details:
            errors += "No capacitor data\n"
            component.details = data_types.Capacitor(value=0, unit=data_types.CapUnits.PF, absolute_pf_value=0,
                                                     voltage=0, tolerance=0, dielectric=list())
        else:
            if not component.details.value:
                errors += "No capacitor value\n"
                component.details.value = 0
            if not component.details.absolute_pf_value:
                component.details.absolute_pf_value = 0
    if component.component_type in [data_types.ComponentType.RESISTOR, data_types.ComponentType.INDUCTOR]:
        type_str: str = data_types.types[component.component_type.value]
        if not component.details:
            errors += 'No %s data\n' % type_str
            component.details = data_types.Resistor(value=-1, tolerance=1)
        elif not component.details.value:
            # we may not have value for inductors
            if component.component_type == data_types.ComponentType.RESISTOR or not component.pn:
                errors += "No %s value\n" % type_str
            component.details.value = -1
    if errors:
        global warning
        warning += 'Component in row %i file %s  is incorrect: ' % (component.row, component.filename)
        warning += errors


def check_for_not_used(row_addr: Row) -> bool:
    """
    checks if component row contains "not used" text
    :param row_addr: information about sheei and row
    :return: true or false
    """
    for i in range(1, row_addr[0].max_column):
        cell = row_addr[0].cell(row=row_addr[1], column=i)
        if cell.value and "not used" in str(cell.value).lower():
            return True
    return False


def get_main_comp_data(row_addr: Row) -> Optional[data_types.Component]:
    """
    gets main component data from xlsx
    :param row_addr:
    :return:
    """
    if check_for_not_used(row_addr):
        return None
    pn: str = get_value('pn', *row_addr) if get_value('pn', *row_addr) else ""
    if pn == "NU":
        return None
    if pn == '-':
        pn = ""
    comp_type: data_types.ComponentType = get_component_type(get_value('type', *row_addr))
    footprint: str = get_footprint_data(get_value('footprint', *row_addr), comp_type)
    pn_alternative: List[str] = [get_value('pn alternative 1', *row_addr), get_value('pn alternative 2', *row_addr)]
    # remove empty alternative pns
    pn_alternative = [x for x in pn_alternative if x]
    designator_raw: str = get_value('designator', *row_addr)
    designator: List[str] = designator_raw.split(', ') if designator_raw else list()
    if not designator and not footprint and not pn:
        return None
    component = data_types.Component(row=row_addr[1], component_type=comp_type, footprint=footprint,
                                     pn=pn, manufacturer=get_value('manufacturer', *row_addr),
                                     designator=designator,
                                     description=get_value('description', *row_addr),
                                     pn_alt=pn_alternative)
    return component


def get_components_from_xlxs(filename) -> List[data_types.Component]:
    """
    parce Bom to get component list
    :param filename: name of BOM
    :return: component list
    """
    global warning
    warning = ""
    wb = load_workbook(filename=filename)
    sheet = wb.active
    result: List[data_types.Component] = list()
    header_index:  Dict[str, Optional[int]] = get_headers(sheet)
    for row in range(2, sheet.max_row):
        if row == 73:
            print('here')
        row_addr: Row = (sheet, row, header_index)
        component: data_types.Component = get_main_comp_data(row_addr)
        if not component:
            warning += "Row %i filename %s not used, skipped\n" % (row, filename)
            continue
        component.filename = os.path.basename(filename)
        if component.component_type == data_types.ComponentType.RESISTOR:
            component.details = get_resistor_data(row_addr)
        elif component.component_type == data_types.ComponentType.CAPACITOR:
            component.details = get_capacitor_data(row_addr)
        elif component.component_type == data_types.ComponentType.INDUCTOR:
            component.details = data_types.Inductor(value=get_value('value', *row_addr))
        validate_and_repair(component)
        result.append(component)
    if warning:
        print('WARNINGS FOR %s:\n' % filename.upper(), warning)
    return result
